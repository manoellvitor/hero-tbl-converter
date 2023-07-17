import openpyxl
from buffer import Buffer
from const import ColType, r_type_titles
from conversions import int_to_bytes
from ctypes import c_uint64


def import_data(in_file):
    wb = openpyxl.load_workbook(in_file)
    sheet = wb.active

    buff = Buffer(bytearray())

    headers = [cell.value for cell in next(sheet.iter_rows())]
    buff.write(str(len(headers)), ColType.UINT32)  # Write column size

    cols = []
    for h in headers:
        col = r_type_titles[h]
        cols.append(col)
        buff.write(str(col.value), ColType.UINT32)  # Write columns

    buff.write("0", ColType.UINT32)  # Write rows size (placeholder)

    for row in sheet.iter_rows(min_row=2):
        for c, cell in enumerate(row):
            col = cols[c]
            value = cell.value
            if col == ColType.STRING:
                value_bytes = value.encode()
                buff.write(str(len(value_bytes)), ColType.UINT32)  # Write text size
                buff.write_bytes(value_bytes)
            else:
                buff.write(str(value), col)

    index = (len(cols) + 1) * 4
    size = buff.get_offset() - c_uint64(index + 8).value

    buff.overwrite(int_to_bytes(sheet.max_row - 1, 4, True), index)  # Write rows size
    buff.overwrite(int_to_bytes(size, 4, True), index + 4)  # Write size

    with open(str(in_file).removesuffix(".xlsx")+ ".tbl", "wb") as file:
        file.write(buff.get_bytes())
        
        return "done"


